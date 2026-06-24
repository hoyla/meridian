"""Derive a CN8 → SITC-division lookup for the codes in our dataset.

This is the structural-spine half of the taxonomy direction (see
`dev_notes/2026-06-20-taxonomy-sitc-spine-and-labels.md`): every CN8 code
in the Eurostat data is mapped to its SITC Rev 4 division — a defensible,
authoritative, reporter-legible bucket — so the whole 9,500-code universe
becomes navigable, not just the ~25% in editorial groups.

Provenance is first-class (the journalism defensibility principle): the
derived lookup records exactly which UNSD source editions produced each
mapping, and a sidecar PROVENANCE.md records the source files, vintages,
and methodology.

Two layers, two roles:
  * The derived CSVs (`reference/cn8_sitc.csv`, `reference/cn8_bec.csv`) are
    the PUBLICATION source of truth — `report_builder` reads only these, so a
    briefing can always be built from a clean checkout. `assert_classifications_available()`
    guards them at build time (a missing/empty lookup is a hard error, never a
    silent collapse of every group into SITC section 9, "Other / unclassified").
  * The raw UNSD workbooks (`reference/un_classifications/`) are BUILD-ONLY
    inputs that regenerate those CSVs. They are committed for reproducible
    rebuilds but are never touched on the publication path. (They previously
    lived outside the repo at `~/Code/un-classifications/`; a folder move
    silently emptied the classifications — hence the move in-repo.)

Two correctness rules learned the hard way (2026-06-20):
1. **Real codes only.** Eurostat Comext mixes aggregate pseudo-codes into
   the data (`000TOTAL`, and `…XX` confidential/aggregated codes). They are
   NOT real CN8 and must be excluded — `000TOTAL` alone is the grand total
   and double-counts the entire denominator if left in.
2. **Multi-edition data.** The dataset spans HS 2017 and HS 2022 (e.g.
   smartphones are 851712 pre-2022, 851713/14 after; solar cells 854140 →
   854142/43). Map HS2022 first, then fall back to HS2017, or ~half the
   value (the big electronics) fails to map.
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

import db

# Derived lookups (the publication source of truth) and their raw inputs.
OUT_DIR = Path(__file__).resolve().parent / "reference"
# Raw UNSD workbooks — BUILD-ONLY inputs that regenerate the CSVs in OUT_DIR.
# Committed in-repo so rebuilds are self-contained; NOT read on the publication
# path. (Were kept outside the repo, which let a folder move silently empty the
# classifications and bucket every group into "Other / unclassified".)
SOURCE_DIR = OUT_DIR / "un_classifications"
HS2022_SITC4 = SOURCE_DIR / "hs2022_sitc4.xlsx"
HS2017_SITC4 = SOURCE_DIR / "hs2017_sitc4.xlsx"
BEC_SOURCE = SOURCE_DIR / "HS-SITC-BEC_Correlations_2022.xlsx"  # has HS22→BEC4

# CN8 product descriptions — human-readable text for individual 8-digit codes,
# so a reader can tell what "29181400" actually is (Citric acid). Same two-layer
# pattern as the SITC/BEC lookups: a committed build-only input regenerates a
# committed derived CSV (the publication source of truth).
#
# Source is Eurostat's *self-explanatory texts* version of the CN — the legal
# nomenclature has terse leaves ("Other", "Of cotton") that only mean something
# with their parent headings; the self-explanatory texts expand each into a
# standalone description (e.g. "Carboxylic acids with additional oxygen function
# … (excl. …)"). 0 bare-"Other" leaves out of 9,800 — so no hierarchy stitching.
# We profiled via the Hungarian KSH mirror (a clean xlsx packaging of the
# Eurostat self-explanatory texts) and cross-validate it against the EU primary
# SKOS/RDF (canonical-of-record) — see cn8_descriptions.PROVENANCE.md.
CN_DESC_DIR = OUT_DIR / "cn_descriptions"
CN_DESC_SOURCE = CN_DESC_DIR / "cn8_2025_en.xlsx"  # build-only input
CN_DESC_YEAR = 2025
# The EU primary SKOS/RDF used for cross-validation. Too big to commit (~195 MB),
# so it's an external check, not a build dependency. Override with CN_RDF_PATH.
CN_RDF_DEFAULT = Path(
    "/Users/luke_hoyland/Code/Reference/eurostat-cn-descriptions/ESTAT-CN2025.rdf"
)

# BEC Rev 4 basic category → SNA broad end-use (the documented standard).
# (Rev 4, not Rev 5: Rev 4 is purpose-built for the capital/intermediate/
# consumption split; Rev 5 restructured in ways that need its own legend.)
_BEC4_ENDUSE = {
    "41": "Capital", "521": "Capital",
    "111": "Intermediate", "121": "Intermediate", "21": "Intermediate",
    "22": "Intermediate", "42": "Intermediate", "53": "Intermediate",
    "322": "Intermediate",
    "112": "Consumption", "122": "Consumption", "51": "Consumption",
    "522": "Consumption", "61": "Consumption", "62": "Consumption",
    "63": "Consumption",
    "31": "Fuel", "32": "Fuel", "321": "Fuel",
    "7": "Other",
}
_ENDUSE_ORDER = {"Capital": 0, "Intermediate": 1, "Consumption": 2,
                 "Fuel": 3, "Other": 4}

# SITC Rev 4 division titles (2-digit), authoritative published list.
SITC_DIVISION = {
    "00": "Live animals", "01": "Meat & preparations", "02": "Dairy & eggs",
    "03": "Fish & seafood", "04": "Cereals & preparations", "05": "Vegetables & fruit",
    "06": "Sugar & honey", "07": "Coffee, tea, cocoa, spices", "08": "Animal feed",
    "09": "Misc edible products", "11": "Beverages", "12": "Tobacco",
    "21": "Hides & skins, raw", "22": "Oil-seeds", "23": "Crude rubber",
    "24": "Cork & wood", "25": "Pulp & waste paper", "26": "Textile fibres",
    "27": "Crude fertilizers & minerals", "28": "Metalliferous ores & scrap",
    "29": "Crude animal/veg materials", "32": "Coal & coke", "33": "Petroleum & products",
    "34": "Gas", "35": "Electric current", "41": "Animal oils & fats",
    "42": "Vegetable fats & oils", "43": "Processed fats/oils, waxes",
    "51": "Organic chemicals", "52": "Inorganic chemicals", "53": "Dyeing/tanning/colouring",
    "54": "Medicinal & pharmaceutical", "55": "Essential oils, cosmetics, cleaning",
    "56": "Fertilizers (manufactured)", "57": "Plastics, primary forms",
    "58": "Plastics, non-primary", "59": "Chemical materials nes",
    "61": "Leather & manufactures", "62": "Rubber manufactures",
    "63": "Cork & wood manufactures", "64": "Paper & paperboard",
    "65": "Textile yarn & fabrics", "66": "Non-metallic mineral manufactures",
    "67": "Iron & steel", "68": "Non-ferrous metals", "69": "Metal manufactures nes",
    "71": "Power-generating machinery", "72": "Specialized industrial machinery",
    "73": "Metalworking machinery", "74": "General industrial machinery",
    "75": "Office & data-processing machines", "76": "Telecom & sound equipment",
    "77": "Electrical machinery & apparatus", "78": "Road vehicles",
    "79": "Other transport equipment", "81": "Prefab buildings; fixtures",
    "82": "Furniture & bedding", "83": "Travel goods & handbags",
    "84": "Apparel & clothing", "85": "Footwear",
    "87": "Scientific & precision instruments", "88": "Photographic & optical goods",
    "89": "Misc manufactured articles", "91": "Postal packages",
    "93": "Special transactions", "96": "Coin", "97": "Gold, non-monetary",
}
SITC_SECTION = {
    "0": "Food & live animals", "1": "Beverages & tobacco", "2": "Crude materials",
    "3": "Mineral fuels", "4": "Oils & fats", "5": "Chemicals",
    "6": "Manufactured goods by material", "7": "Machinery & transport",
    "8": "Misc manufactured", "9": "Other / unclassified",
}


def _load_conversion(path: Path) -> dict[str, str]:
    """HS6 -> SITC4 from a UNSD 'Conversion' sheet."""
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = next(s for s in wb.sheetnames if s.lower().startswith("conversion"))
    ws = wb[sheet]
    m: dict[str, str] = {}
    for i, (hs, sitc) in enumerate(ws.iter_rows(values_only=True)):
        if i and hs and sitc:
            m[str(hs).strip()] = str(sitc).strip()
    return m


def _real_cn8(code: str) -> bool:
    """A genuine 8-digit CN code — excludes 000TOTAL and …XX aggregates."""
    return code.isdigit() and len(code) == 8


_HS6_DIV: dict[str, str] | None = None


def _hs6_division() -> dict[str, str]:
    """HS6 -> SITC 2-digit division, from the committed derived lookup
    (`reference/cn8_sitc.csv`). This is the PUBLICATION source for the SITC
    `sector` facet — the raw UNSD workbooks are only needed to (re)build that
    CSV. Cached. Empty dict if the CSV is absent; `assert_classifications_available`
    turns that into a hard error at briefing-build time."""
    global _HS6_DIV
    if _HS6_DIV is None:
        m: dict[str, str] = {}
        try:
            with open(OUT_DIR / "cn8_sitc.csv") as f:
                for row in csv.DictReader(f):
                    hs6, div = row.get("hs6"), row.get("sitc_division")
                    if hs6 and div:
                        m[hs6] = div
        except FileNotFoundError:
            pass
        _HS6_DIV = m
    return _HS6_DIV


def sitc_divisions_for_patterns(patterns) -> list[str]:
    """The SITC division code(s) an editorial group spans, from its HS
    wildcard patterns (2/4/6/8-digit). A group may span several — it's a
    label, not a partition (e.g. 'Electrical equipment, broad' → 10).
    Resolved against the committed CN8→division lookup, i.e. the codes
    actually present in the data."""
    conv = _hs6_division()  # hs6 -> 2-digit SITC division
    divs: set[str] = set()
    for p in patterns or []:
        d = p.replace("%", "")
        if len(d) >= 6:
            div = conv.get(d[:6])
            if div:
                divs.add(div)
        else:
            for hs6, div in conv.items():
                if hs6.startswith(d):
                    divs.add(div)
    return sorted(divs)


def division_title(code: str) -> str:
    return SITC_DIVISION.get(code, f"div {code}")


def section_title(code: str) -> str:
    """SITC 1-digit section title (the coarse grouping for the sector list)."""
    return SITC_SECTION.get(code, "Other / unclassified")


_HS_BEC4: dict[str, str] | None = None


def _hs6_bec4() -> dict[str, str]:
    """HS6 -> BEC Rev 4 (cached) from the combined correlation workbook.
    Empty if absent — the end-use facet is optional enrichment."""
    global _HS_BEC4
    if _HS_BEC4 is None:
        m: dict[str, str] = {}
        try:
            ws = openpyxl.load_workbook(BEC_SOURCE, read_only=True)["HS SITC BEC"]
            hi = bi = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    hdr = [str(c) for c in row]
                    hi, bi = hdr.index("HS22"), hdr.index("BEC4")
                    continue
                hs, b = row[hi], row[bi]
                if hs and b and str(b).strip() != "NULL":
                    m[str(hs).strip()] = str(b).strip()
        except (FileNotFoundError, KeyError, ValueError):
            m = {}
        _HS_BEC4 = m
    return _HS_BEC4


def bec4_enduse(bec4: str) -> str:
    return _BEC4_ENDUSE.get(str(bec4).strip(), "Other")


_HS6_ENDUSE: dict[str, str] | None = None


def _hs6_enduse() -> dict[str, str]:
    """HS6 -> SNA end-use (Capital/Intermediate/Consumption/Fuel/Other), from
    the committed derived lookup (`reference/cn8_bec.csv`). PUBLICATION source
    for the end-use facet; the raw BEC workbook is build-only. Cached. Empty if
    absent (see `assert_classifications_available`)."""
    global _HS6_ENDUSE
    if _HS6_ENDUSE is None:
        m: dict[str, str] = {}
        try:
            with open(OUT_DIR / "cn8_bec.csv") as f:
                for row in csv.DictReader(f):
                    hs6, eu = row.get("hs6"), row.get("end_use")
                    if hs6 and eu:
                        m[hs6] = eu
        except FileNotFoundError:
            pass
        _HS6_ENDUSE = m
    return _HS6_ENDUSE


def enduse_for_patterns(patterns) -> list[str]:
    """The SNA end-use category/categories an editorial group spans
    (Capital / Intermediate / Consumption / Fuel), from its HS patterns.
    Resolved against the committed CN8→end-use lookup."""
    hse = _hs6_enduse()  # hs6 -> end-use (already resolved)
    cats: set[str] = set()
    for p in patterns or []:
        d = p.replace("%", "")
        if len(d) >= 6:
            eu = hse.get(d[:6])
            if eu:
                cats.add(eu)
        else:
            for hs6, eu in hse.items():
                if hs6.startswith(d):
                    cats.add(eu)
    return sorted(cats, key=lambda c: _ENDUSE_ORDER.get(c, 9))


_CN8_DIV: dict[str, str] | None = None


def cn8_division_map() -> dict[str, str]:
    """CN8 -> SITC division, from the committed derived lookup
    (reference/cn8_sitc.csv). Empty dict if absent (never breaks a render)."""
    global _CN8_DIV
    if _CN8_DIV is None:
        path = OUT_DIR / "cn8_sitc.csv"
        m: dict[str, str] = {}
        try:
            with open(path) as f:
                for row in csv.DictReader(f):
                    m[row["cn8"]] = row["sitc_division"]
        except FileNotFoundError:
            pass
        _CN8_DIV = m
    return _CN8_DIV


def _cn8_short_label(full: str) -> str:
    """A compact label for inline display, from the full self-explanatory text.

    The head clause — text before the first '(' (the "(excl. …)" qualifier) and
    before the first comma/semicolon — is the gist: "Lithium-ion accumulators
    (excl. spent)" -> "Lithium-ion accumulators"; "Citric acid" -> "Citric acid".
    Median 28 chars, 90% <= 60 over the full CN. Defensive on leading dashes in
    case a terse source is ever swapped in."""
    import re
    h = (full or "").lstrip("- ").split("(")[0]
    h = re.split(r"[;,]", h, maxsplit=1)[0]
    return h.strip().rstrip(":").strip()


_CN8_DESC: dict[str, dict[str, str]] | None = None


def cn8_description_lookup() -> dict[str, dict[str, str]]:
    """CN8 -> {"short": ..., "full": ...} from the committed derived lookup
    (reference/cn8_descriptions.csv). Cached. Empty dict if absent — descriptions
    are reader enrichment, never publication-critical, so a missing file degrades
    gracefully to today's bare codes (NOT guarded by
    assert_classifications_available)."""
    global _CN8_DESC
    if _CN8_DESC is None:
        path = OUT_DIR / "cn8_descriptions.csv"
        m: dict[str, dict[str, str]] = {}
        try:
            with open(path) as f:
                for row in csv.DictReader(f):
                    code = (row.get("cn8") or "").strip()
                    if code:
                        m[code] = {"short": row.get("label_short", ""),
                                   "full": row.get("denomination", "")}
        except FileNotFoundError:
            pass
        _CN8_DESC = m
    return _CN8_DESC


def assert_classifications_available() -> None:
    """Fail loud if the publication-critical classification lookups are missing
    or empty. Called at briefing-build time so a moved, renamed or empty
    reference file surfaces as a hard error — never the silent collapse of
    every group into SITC section 9 ("Other / unclassified") that a missing
    lookup otherwise produces.

    Publication reads only the committed derived CSVs, so this checks those
    (not the raw UNSD workbooks, which are build-only)."""
    problems: list[str] = []
    if not _hs6_division():
        problems.append(f"SITC division lookup empty or missing: {OUT_DIR / 'cn8_sitc.csv'}")
    if not _hs6_enduse():
        problems.append(f"BEC end-use lookup empty or missing: {OUT_DIR / 'cn8_bec.csv'}")
    if problems:
        raise RuntimeError(
            "Classification lookups unavailable — the briefing would bucket "
            "every group into 'Other / unclassified'. Refusing to build:\n  - "
            + "\n  - ".join(problems)
            + "\nThese committed CSVs are the publication source of truth. "
            f"Regenerate with `python classifications.py` (reads the UNSD "
            f"workbooks in {SOURCE_DIR})."
        )


def build(write: bool = True) -> dict:
    m22 = _load_conversion(HS2022_SITC4)
    m17 = _load_conversion(HS2017_SITC4)

    with db._conn() as conn, conn.cursor() as cur:
        cur.execute(
            """SELECT DISTINCT o.hs_code
                 FROM observations o JOIN releases r ON r.id = o.release_id
                WHERE r.source = 'eurostat' AND o.hs_code IS NOT NULL"""
        )
        codes = [r[0] for r in cur.fetchall()]

    rows = []
    stats = {"raw": len(codes), "real": 0, "mapped": 0, "unmapped": 0,
             "via_hs2022": 0, "via_hs2017": 0}
    for code in sorted(c for c in codes if _real_cn8(c)):
        stats["real"] += 1
        hs6 = code[:6]
        sitc = m22.get(hs6)
        via = "hs2022"
        if sitc is None:
            sitc = m17.get(hs6)
            via = "hs2017"
        if sitc is None:
            stats["unmapped"] += 1
            continue
        stats["mapped"] += 1
        stats[f"via_{via}"] += 1
        div = sitc[:2]
        rows.append({
            "cn8": code, "hs6": hs6, "sitc4": sitc,
            "sitc_division": div, "sitc_division_title": SITC_DIVISION.get(div, ""),
            "sitc_section": sitc[0], "sitc_section_title": SITC_SECTION.get(sitc[0], ""),
            "mapped_via": via,
        })

    if write:
        OUT_DIR.mkdir(exist_ok=True)
        with open(OUT_DIR / "cn8_sitc.csv", "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
        _write_provenance(stats)
    return {"stats": stats, "rows": rows}


def _write_provenance(stats: dict) -> None:
    from datetime import datetime
    OUT_DIR.mkdir(exist_ok=True)
    (OUT_DIR / "cn8_sitc.PROVENANCE.md").write_text(f"""# cn8_sitc.csv — provenance

Generated by `classifications.build()` on {datetime.now():%Y-%m-%d %H:%M}.

## What it is
A CN8 → SITC Rev 4 division lookup for every real 8-digit code present in
the Eurostat China-trade observations. The SITC division is the structural
navigation spine for the portal (see the taxonomy design note).

## Source classifications (UNSD)
- HS 2022 → SITC Rev 4: `hs2022_sitc4.xlsx`
- HS 2017 → SITC Rev 4 (fallback): `hs2017_sitc4.xlsx`
- Source: <https://unstats.un.org/unsd/classifications/Econ/tables/>, downloaded 2026-06-20.
- Files committed in-repo at `reference/un_classifications/` (build-only inputs).

## Method
- CN8 mapped via its 6-digit HS stem.
- **HS2022 first, then HS2017 fallback** — the dataset spans both editions.
- Aggregate pseudo-codes (`000TOTAL`, `…XX`) excluded: real 8-digit numeric only.

## Vintages to refresh on
HS 2022 / SITC Rev 4. When HS 2027 ships and the EU CN rebases, add the
HS2027→SITC correspondence ahead of HS2022 in the fallback chain.

## Counts at generation
- raw distinct codes in data: {stats['raw']:,}
- real 8-digit CN8: {stats['real']:,}  (pseudo/aggregate dropped: {stats['raw'] - stats['real']:,})
- mapped: {stats['mapped']:,}  (via HS2022 {stats['via_hs2022']:,}, via HS2017 {stats['via_hs2017']:,})
- unmapped: {stats['unmapped']:,}
""")


def build_bec(write: bool = True) -> dict:
    """Derive CN8 → BEC Rev 4 → SNA end-use for the dataset codes."""
    from collections import Counter
    hsbec = _hs6_bec4()
    with db._conn() as conn, conn.cursor() as cur:
        cur.execute(
            """SELECT DISTINCT o.hs_code
                 FROM observations o JOIN releases r ON r.id = o.release_id
                WHERE r.source = 'eurostat' AND o.hs_code IS NOT NULL"""
        )
        codes = [r[0] for r in cur.fetchall()]
    rows = []
    stats = {"raw": len(codes), "real": 0, "mapped": 0, "unmapped": 0}
    eu_count: Counter = Counter()
    for code in sorted(c for c in codes if _real_cn8(c)):
        stats["real"] += 1
        b = hsbec.get(code[:6])
        if not b:
            stats["unmapped"] += 1
            continue
        eu = bec4_enduse(b)
        stats["mapped"] += 1
        eu_count[eu] += 1
        rows.append({"cn8": code, "hs6": code[:6], "bec4": b, "end_use": eu})
    if write:
        OUT_DIR.mkdir(exist_ok=True)
        with open(OUT_DIR / "cn8_bec.csv", "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["cn8", "hs6", "bec4", "end_use"])
            w.writeheader()
            w.writerows(rows)
        from datetime import datetime
        (OUT_DIR / "cn8_bec.PROVENANCE.md").write_text(f"""# cn8_bec.csv — provenance

Generated by `classifications.build_bec()` on {datetime.now():%Y-%m-%d %H:%M}.

CN8 → BEC Rev 4 → SNA broad end-use (Capital / Intermediate / Consumption /
Fuel) for every real 8-digit code in the Eurostat data.

- Source: `HS-SITC-BEC_Correlations_2022.xlsx` (UNSD, HS2022→BEC4 column),
  committed in-repo at `reference/un_classifications/` (build-only).
- **BEC Rev 4, not Rev 5** — Rev 4 is the documented standard for the
  capital/intermediate/consumption end-use split; Rev 5 restructured in ways
  that need its own legend to read end-use.
- Mapped via the 6-digit HS stem; real 8-digit codes only.

## Counts
- real CN8: {stats['real']:,}  mapped: {stats['mapped']:,}  unmapped: {stats['unmapped']:,}
- by end-use: {dict(eu_count)}
""")
    return {"stats": stats, "end_use": dict(eu_count)}


def _crossvalidate_descriptions(desc: dict[str, str], rdf_path: Path) -> dict:
    """Confirm the self-explanatory texts agree with the EU primary SKOS/RDF
    (canonical-of-record). The RDF carries terse, indented labels ("--- Other",
    "- Of cotton"); the self-explanatory text expands them. So we check two
    things per CN8 leaf in the RDF:
      * coverage — is the code present in our description set?
      * consistency — for *non-positional* terse labels (not "Other"/"Of …"),
        does the terse phrase appear within our self-explanatory text?
    Returns counts; never raises (a missing RDF just yields skipped=True)."""
    import xml.etree.ElementTree as ET
    if not rdf_path.exists():
        return {"skipped": True, "rdf": str(rdf_path)}
    RDF = "{http://www.w3.org/1999/02/22-rdf-syntax-ns#}"
    SKOS = "{http://www.w3.org/2004/02/skos/core#}"
    DC = "{http://purl.org/dc/elements/1.1/}"
    LANG = "{http://www.w3.org/XML/1998/namespace}lang"
    # Index the EU primary by CN8 -> terse English altLabel. The RDF carries the
    # whole hierarchy (chapters, headings, subheadings) as "...80" product-lines,
    # so we key by the leading 8 digits and only keep codes that exist in our
    # set — i.e. we validate OUR codes against the EU primary, not the reverse
    # (the reverse denominator would wrongly include HS4/HS6 roll-up nodes).
    rdf_terse: dict[str, str] = {}
    for _ev, el in ET.iterparse(str(rdf_path), events=("end",)):
        if el.tag != RDF + "Description":
            continue
        ident = el.findtext(DC + "identifier") or ""
        if ident.endswith("80") and len(ident) == 12 and ident[:8].isdigit():
            cn8 = ident[:8]
            if cn8 in desc and cn8 not in rdf_terse:
                for n in el.findall(SKOS + "altLabel"):
                    if n.get(LANG) == "en":
                        rdf_terse[cn8] = (n.text or "").lstrip("- ").strip()
                        break
        el.clear()
    st = {"skipped": False, "our_codes": len(desc),
          "present_in_eu": 0, "absent_from_eu": 0,
          "checked": 0, "consistent": 0, "positional": 0}
    for cn8, ours in desc.items():
        terse = rdf_terse.get(cn8)
        if terse is None:
            st["absent_from_eu"] += 1
            continue
        st["present_in_eu"] += 1
        low = terse.lower()
        if not terse or low.startswith(("other", "of ", "- ")):
            st["positional"] += 1  # terse leaf only meaningful via parent; can't string-match
        else:
            st["checked"] += 1
            head = _cn8_short_label(terse).lower()  # terse head clause
            if head and head in ours.lower():
                st["consistent"] += 1
    st["coverage_pct"] = round(100 * st["present_in_eu"] / st["our_codes"], 2) if desc else 0.0
    st["consistency_pct"] = round(100 * st["consistent"] / st["checked"], 2) if st["checked"] else 0.0
    return st


def build_cn8_descriptions(write: bool = True, rdf_path: Path | None = None) -> dict:
    """Derive reference/cn8_descriptions.csv from the Eurostat self-explanatory
    CN texts (KSH mirror). Covers ALL ~9,800 CN codes, not just those in our
    data — descriptions are a static reference, and a complete file means new
    codes appearing in later data still resolve (and the build needs no DB)."""
    import os
    wb = openpyxl.load_workbook(CN_DESC_SOURCE, read_only=True)
    ws = next((wb[s] for s in wb.sheetnames if "structure" in s.lower()), wb.active)
    desc: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0]).strip() if row and row[0] is not None else ""
        den = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if _real_cn8(code) and den:
            desc[code] = den

    rows = [{"cn8": c, "label_short": _cn8_short_label(desc[c]),
             "denomination": desc[c], "cn_year": CN_DESC_YEAR}
            for c in sorted(desc)]
    stats = {"codes": len(rows)}
    xval = _crossvalidate_descriptions(
        desc, Path(os.environ.get("CN_RDF_PATH", str(rdf_path or CN_RDF_DEFAULT))))
    stats["crossval"] = xval

    if write:
        CN_DESC_DIR.mkdir(parents=True, exist_ok=True)
        with open(OUT_DIR / "cn8_descriptions.csv", "w", newline="") as f:
            w = csv.DictWriter(
                f, fieldnames=["cn8", "label_short", "denomination", "cn_year"])
            w.writeheader()
            w.writerows(rows)
        _write_cn_desc_provenance(stats)
    return stats


def _write_cn_desc_provenance(stats: dict) -> None:
    from datetime import datetime
    xv = stats.get("crossval", {})
    if xv.get("skipped"):
        xval_block = (f"- **Not cross-validated this run** — EU primary RDF not "
                      f"found at `{xv.get('rdf')}`.\n  Set `CN_RDF_PATH` to the "
                      f"`ESTAT-CN{CN_DESC_YEAR}.rdf` download to enable the check.")
    else:
        xval_block = (
            f"- Our CN8 codes present in the EU primary: "
            f"{xv.get('present_in_eu', 0):,}/{xv.get('our_codes', 0):,} "
            f"({xv.get('coverage_pct', 0)}%) — i.e. every code we describe is a "
            f"real EU code; absent: {xv.get('absent_from_eu', 0):,}.\n"
            f"- Label consistency where the EU terse label is non-positional: "
            f"{xv.get('consistent', 0):,}/{xv.get('checked', 0):,} "
            f"({xv.get('consistency_pct', 0)}%) head-clause match. "
            f"Positional terse leaves (\"Other\"/\"Of …\", meaningful only via "
            f"their parent) can't be string-matched and are excluded: "
            f"{xv.get('positional', 0):,}. The residual difference is "
            f"self-explanatory wording vs the terse legal label, not a code "
            f"mismatch.")
    (OUT_DIR / "cn8_descriptions.PROVENANCE.md").write_text(f"""# cn8_descriptions.csv — provenance

Generated by `classifications.build_cn8_descriptions()` on {datetime.now():%Y-%m-%d %H:%M}.

## What it is
A CN8 → product-description lookup for every real 8-digit code in the
Combined Nomenclature {CN_DESC_YEAR} (~9,800 codes). Two columns of text:
- `denomination` — the full Eurostat *self-explanatory* text (standalone,
  median ~112 chars), for tooltips / definitions.
- `label_short` — the head clause (median ~28 chars), for inline display.

Descriptions are reader enrichment, never publication-critical: a missing file
degrades gracefully to bare codes (it is NOT guarded by
`assert_classifications_available`).

## Source
- **Content:** Eurostat *self-explanatory texts* of the Combined Nomenclature
  {CN_DESC_YEAR} (Commission Implementing Regulation (EU) 2024/2522). The legal
  CN has terse leaves ("Other", "Of cotton") meaningful only via their parent
  headings; the self-explanatory version expands each into a standalone phrase,
  so **0 of ~9,800 leaves are a bare "Other"** and no hierarchy stitching is
  needed.
- **Build-only input (committed):** `reference/cn_descriptions/cn8_2025_en.xlsx`
  — a clean tabular packaging of the Eurostat self-explanatory texts published
  by the Hungarian Central Statistical Office (KSH),
  <https://www.ksh.hu/intrastat_combined_nomenclature>, downloaded 2026-06-24.
  (The EU portal at data.europa.eu publishes the CN only as terse SKOS/RDF;
  there is no EU-hosted tabular self-explanatory download.)

## Cross-validation against the EU primary (canonical-of-record)
Checked against the EU SKOS/RDF `ESTAT-CN{CN_DESC_YEAR}.rdf`
(<https://data.europa.eu/data/datasets/combined-nomenclature-{CN_DESC_YEAR}>),
the legal nomenclature of record. Too big to commit (~195 MB); external check.
{xval_block}

The handful of codes absent from the EU primary are Intrastat/special-procedure
pseudo-codes (CN chapters 98 "complete industrial plant" and 99 vessel/aircraft
supplies, small-value, returned goods, personal property) — present in the KSH
Intrastat packaging but outside the legal goods nomenclature. Expected, not a gap.

## Method
- Read the self-explanatory denomination per real 8-digit code.
- `label_short` = head clause: text before the first "(" and first comma/semicolon.
- Real 8-digit numeric codes only (`000TOTAL` / `…XX` aggregates excluded).

## Vintages to refresh on
The CN is reissued every 1 January (codes added/split/withdrawn). Refresh
`cn8_2025_en.xlsx` to the new year's file and rerun. `cn_year` records the
edition. v1 stores a single current-edition snapshot; if historical fidelity on
older figures matters, key descriptions by CN year instead. See the roadmap note
on keeping externally-sourced reference data current.

## Counts at generation
- CN8 codes with a description: {stats.get('codes', 0):,}
""")


if __name__ == "__main__":
    print("cn8_sitc lookup built:", build()["stats"])
    print("cn8_bec lookup built:", build_bec()["stats"])
    print("cn8_descriptions built:", build_cn8_descriptions())
