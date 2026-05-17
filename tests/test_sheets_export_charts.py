"""Tests for the xlsx Charts tab.

Companion to `test_briefing_pack_docx.py` — the Charts tab is the
spreadsheet counterpart to the docx output's per-mover chart cards.
Same input (top-N movers + their monthly eurostat_raw_rows series),
different output format.

Approach: seed findings + raw rows, run XlsxWriter with charts=True,
load result back with openpyxl, assert presence + count + layout of
native chart objects and the per-mover data blocks they reference.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import psycopg2
import pytest
from openpyxl import load_workbook

import sheets_export
from tests.test_briefing_pack import _seed_eurostat_release, _seed_run
from tests.test_briefing_pack_docx import (
    _seed_chart_capable_finding,
    _seed_eurostat_raw_rows_for_finding,
)


@pytest.fixture(autouse=True)
def _direct_db_url(test_db_url, monkeypatch):
    monkeypatch.setenv("DATABASE_URL", test_db_url)
    monkeypatch.delenv(sheets_export.PERMALINK_BASE_ENV, raising=False)


@pytest.fixture
def empty_findings(test_db_url):
    with psycopg2.connect(test_db_url) as conn, conn.cursor() as cur:
        cur.execute(
            "TRUNCATE findings, observations, source_snapshots, "
            "eurostat_raw_rows, scrape_runs, releases "
            "RESTART IDENTITY CASCADE"
        )
    yield


def _seed_one_mover(test_db_url):
    """Seed a single eligible mover + its underlying raw rows."""
    with psycopg2.connect(test_db_url) as conn:
        cur = conn.cursor()
        run = _seed_run(cur)
        _seed_eurostat_release(cur, date(2026, 2, 1))
        _seed_chart_capable_finding(
            cur, run, "EV batteries (Li-ion)",
            yoy_pct=0.35, current_eur=27e9, prior_eur=20e9, low_base=False,
        )
        _seed_eurostat_raw_rows_for_finding(cur, run, hs_pattern="8507%")
        conn.commit()


def _seed_three_movers(test_db_url):
    """Seed three eligible movers for top_n truncation testing."""
    with psycopg2.connect(test_db_url) as conn:
        cur = conn.cursor()
        run = _seed_run(cur)
        _seed_eurostat_release(cur, date(2026, 2, 1))
        _seed_chart_capable_finding(
            cur, run, "EV batteries (Li-ion)",
            yoy_pct=0.35, current_eur=27e9, prior_eur=20e9,
        )
        _seed_chart_capable_finding(
            cur, run, "Drones and unmanned aircraft",
            yoy_pct=0.40, current_eur=1.0e9, prior_eur=0.7e9,
        )
        _seed_chart_capable_finding(
            cur, run, "Wind generating sets only",
            yoy_pct=0.345, current_eur=375e6, prior_eur=279e6,
        )
        _seed_eurostat_raw_rows_for_finding(cur, run, hs_pattern="8507%")
        conn.commit()


def test_charts_disabled_by_default(empty_findings, test_db_url, tmp_path):
    """Without `charts=True`, no Charts tab is created — back-compat
    guard so existing pipelines (cycles without `--docx`) get the same
    workbook as before."""
    _seed_one_mover(test_db_url)
    out = tmp_path / "no-charts.xlsx"

    sheets = sheets_export.assemble_sheets()
    sheets_export.XlsxWriter().write(sheets, str(out))

    wb = load_workbook(str(out))
    assert "Charts" not in wb.sheetnames


def test_charts_tab_created_when_enabled(
    empty_findings, test_db_url, tmp_path,
):
    """`charts=True` adds a Charts tab at the end."""
    _seed_one_mover(test_db_url)
    out = tmp_path / "with-charts.xlsx"

    sheets = sheets_export.assemble_sheets()
    sheets_export.XlsxWriter().write(sheets, str(out), charts=True)

    wb = load_workbook(str(out))
    assert "Charts" in wb.sheetnames
    # Charts tab is last (added after regular SheetData tabs)
    assert wb.sheetnames[-1] == "Charts"


def test_charts_tab_contains_native_line_chart_per_mover(
    empty_findings, test_db_url, tmp_path,
):
    """N movers → N native openpyxl LineChart objects on the Charts tab.
    Native chart (not flat image) is the property that survives upload
    to Google Sheets as an editable chart."""
    _seed_three_movers(test_db_url)
    out = tmp_path / "native-charts.xlsx"

    sheets_export.XlsxWriter().write(
        sheets_export.assemble_sheets(), str(out),
        charts=True, charts_top_n=3,
    )

    wb = load_workbook(str(out))
    charts_ws = wb["Charts"]
    # `_charts` is openpyxl's internal list of chart objects on the sheet
    assert len(charts_ws._charts) == 3


def test_charts_tab_respects_top_n(
    empty_findings, test_db_url, tmp_path,
):
    """Seeding 3 movers but requesting top_n=2 → only 2 mover blocks."""
    _seed_three_movers(test_db_url)
    out = tmp_path / "topn-charts.xlsx"

    sheets_export.XlsxWriter().write(
        sheets_export.assemble_sheets(), str(out),
        charts=True, charts_top_n=2,
    )

    wb = load_workbook(str(out))
    charts_ws = wb["Charts"]
    assert len(charts_ws._charts) == 2


def test_charts_tab_data_layout(
    empty_findings, test_db_url, tmp_path,
):
    """Per-mover data block has the right shape:
    - Heading row (group name + flow label)
    - Column header row (Month / Prior 12mo (€) / Current 12mo (€))
    - 24 data rows with prior populating first 12 and current the next 12

    Tests the layout invariant that LineChart references count on."""
    _seed_one_mover(test_db_url)
    out = tmp_path / "layout.xlsx"
    sheets_export.XlsxWriter().write(
        sheets_export.assemble_sheets(), str(out),
        charts=True, charts_top_n=1,
    )

    wb = load_workbook(str(out))
    ws = wb["Charts"]

    # Row 3: heading. Row 4: column headers.
    assert "EV batteries (Li-ion)" in (ws.cell(row=3, column=1).value or "")
    assert ws.cell(row=4, column=1).value == "Month"
    assert ws.cell(row=4, column=2).value == "Prior 12mo (€)"
    assert ws.cell(row=4, column=3).value == "Current 12mo (€)"

    # Rows 5-16 are prior (col B populated, col C None).
    # Rows 17-28 are current (col B None, col C populated).
    prior_populated = sum(
        1 for r in range(5, 17)
        if ws.cell(row=r, column=2).value is not None
    )
    current_populated = sum(
        1 for r in range(17, 29)
        if ws.cell(row=r, column=3).value is not None
    )
    assert prior_populated == 12
    assert current_populated == 12

    # Cross-checks: no current values in the prior rows, no prior in current.
    for r in range(5, 17):
        assert ws.cell(row=r, column=3).value is None, f"row {r} has current value"
    for r in range(17, 29):
        assert ws.cell(row=r, column=2).value is None, f"row {r} has prior value"


def test_charts_tab_unavailable_branch(
    empty_findings, test_db_url, tmp_path,
):
    """A mover with no underlying raw rows → "Chart unavailable" italic
    line in the block, and no LineChart for that mover."""
    # Seed the finding WITHOUT raw rows
    with psycopg2.connect(test_db_url) as conn:
        cur = conn.cursor()
        run = _seed_run(cur)
        _seed_eurostat_release(cur, date(2026, 2, 1))
        _seed_chart_capable_finding(
            cur, run, "EV batteries (Li-ion)",
            yoy_pct=0.35, current_eur=27e9, prior_eur=20e9,
        )
        conn.commit()
    out = tmp_path / "unavailable.xlsx"
    sheets_export.XlsxWriter().write(
        sheets_export.assemble_sheets(), str(out),
        charts=True, charts_top_n=1,
    )

    wb = load_workbook(str(out))
    ws = wb["Charts"]
    # No LineChart objects (the unavailable branch skips chart insertion)
    assert len(ws._charts) == 0
    # The "Chart unavailable" marker is present
    found_marker = False
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and "Chart unavailable" in str(cell):
                found_marker = True
                break
    assert found_marker


def test_charts_tab_description_row(
    empty_findings, test_db_url, tmp_path,
):
    """The tab opens with an italic description row explaining what's
    inside — same convention as the other SheetData tabs."""
    _seed_one_mover(test_db_url)
    out = tmp_path / "description.xlsx"
    sheets_export.XlsxWriter().write(
        sheets_export.assemble_sheets(), str(out), charts=True,
    )

    wb = load_workbook(str(out))
    ws = wb["Charts"]
    description = ws.cell(row=1, column=1).value or ""
    assert "monthly" in description.lower()
    assert "03_findings.docx" in description.lower()
