"""Tests for the Sheets/XLSX exporter.

Approach: seed a few findings, run the exporter to a temp .xlsx, then load
it back with openpyxl and assert structure + values. We don't mock the SQL
layer — the queries are part of what we want to test.
"""

import json
from datetime import date
from pathlib import Path

import psycopg2
import pytest
from openpyxl import load_workbook

import sheets_export


@pytest.fixture(autouse=True)
def _direct_db_url(test_db_url, monkeypatch):
    monkeypatch.setenv("DATABASE_URL", test_db_url)
    # Ensure permalink base is unset so the `link` column stays empty for tests
    # that don't explicitly set it.
    monkeypatch.delenv(sheets_export.PERMALINK_BASE_ENV, raising=False)


@pytest.fixture
def empty_findings(test_db_url):
    with psycopg2.connect(test_db_url) as conn, conn.cursor() as cur:
        cur.execute(
            "TRUNCATE findings, observations, source_snapshots, eurostat_raw_rows, "
            "scrape_runs, releases RESTART IDENTITY CASCADE"
        )
    yield


def _seed_one_finding(conn, subkind: str, group_name: str, detail: dict) -> int:
    cur = conn.cursor()
    cur.execute("INSERT INTO scrape_runs (source_url, status) VALUES ('seed', 'success') RETURNING id")
    run = cur.fetchone()[0]
    cur.execute(
        "SELECT id FROM hs_groups WHERE name = %s", (group_name,),
    )
    hg = cur.fetchone()
    hg_ids = [hg[0]] if hg else []
    cur.execute(
        """
        INSERT INTO findings (scrape_run_id, kind, subkind, observation_ids, hs_group_ids,
                              score, title, body, detail)
        VALUES (%s, 'anomaly', %s, '{}', %s, %s, %s, 'b', %s::jsonb)
        RETURNING id
        """,
        (run, subkind, hg_ids, abs(detail.get("totals", {}).get("yoy_pct", 0.0)),
         f"seed {subkind} {group_name}", json.dumps(detail)),
    )
    fid = cur.fetchone()[0]
    conn.commit()
    return fid


def test_export_produces_xlsx_with_all_tabs(empty_findings, test_db_url, tmp_path):
    """An empty findings table still produces all 8 tabs — just with no data rows.
    Tab roster is documented in sheets_export.assemble_sheets()."""
    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    expected_tabs = {
        "summary",
        "hs_yoy_imports", "hs_yoy_exports",
        "trajectories",
        "mirror_gaps", "mirror_gap_movers",
        "low_base_review",
        "predictability_index",
    }
    assert set(wb.sheetnames) == expected_tabs


def test_summary_picks_latest_per_group_per_scope(empty_findings, test_db_url, tmp_path):
    """The summary sheet's per-(scope, flow) columns show the *latest* finding
    per group, not duplicates and not the older one."""
    period_old = date(2025, 12, 1)
    period_new = date(2026, 2, 1)
    with psycopg2.connect(test_db_url) as conn:
        for period, yoy in [(period_old, 0.10), (period_new, 0.40)]:
            _seed_one_finding(conn, "hs_group_yoy", "EV batteries (Li-ion)", {
                "windows": {"current_end": period.isoformat(), "current_start": period.isoformat()},
                "totals": {
                    "yoy_pct": yoy, "current_12mo_eur": 1e9, "prior_12mo_eur": 0.9e9,
                    "yoy_pct_kg": yoy * 1.5, "current_12mo_kg": 1e6,
                    "low_base": False,
                },
                "group": {"name": "EV batteries (Li-ion)"},
            })

    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["summary"]
    # Row 1 is the description, row 3 is headers, row 4+ is data.
    headers = [c.value for c in ws[3]]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=4, values_only=False)]
    ev_rows = [r for r in rows if r[headers.index("group")] == "EV batteries (Li-ion)"]
    assert len(ev_rows) == 1
    # eu_27 imports column carries the latest yoy_pct
    assert ev_rows[0][headers.index("eu27_imp_yoy_pct")] == 0.40


def test_hs_yoy_imports_long_format_with_scope_column(
    empty_findings, test_db_url, tmp_path,
):
    """The hs_yoy_imports tab is long-format: one row per (group, scope).
    Scope column lets a journalist filter."""
    with psycopg2.connect(test_db_url) as conn:
        # Seed three findings: same group, three different scope subkinds
        for subkind, expected_scope, yoy in [
            ("hs_group_yoy", "eu_27", 0.40),
            ("hs_group_yoy_uk", "uk", 0.20),
            ("hs_group_yoy_combined", "eu_27_plus_uk", 0.30),
        ]:
            _seed_one_finding(conn, subkind, "EV batteries (Li-ion)", {
                "windows": {"current_end": "2026-02-01", "current_start": "2026-02-01"},
                "totals": {"yoy_pct": yoy, "current_12mo_eur": 1e9, "prior_12mo_eur": 0.9e9,
                           "low_base": False, "low_base_threshold_eur": 5e7},
                "group": {"name": "EV batteries (Li-ion)"},
            })

    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["hs_yoy_imports"]
    headers = [c.value for c in ws[3]]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=4)]
    ev_rows = [r for r in rows if r[headers.index("group")] == "EV batteries (Li-ion)"]
    # All three scopes should appear
    scopes = {r[headers.index("scope")] for r in ev_rows}
    assert scopes == {"eu_27", "uk", "eu_27_plus_uk"}


def test_hs_yoy_includes_threshold_fragility_flag(
    empty_findings, test_db_url, tmp_path,
):
    """Findings within 1.5x of the low_base threshold should have
    near_low_base_threshold = True even when low_base = False."""
    with psycopg2.connect(test_db_url) as conn:
        # current=€55M, prior=€60M: above threshold (low_base=False) but
        # within 1.5x the €50M threshold (near_low_base_threshold=True)
        _seed_one_finding(conn, "hs_group_yoy", "EV batteries (Li-ion)", {
            "windows": {"current_end": "2026-02-01", "current_start": "2026-02-01"},
            "totals": {
                "yoy_pct": 0.10, "current_12mo_eur": 5.5e7, "prior_12mo_eur": 6.0e7,
                "low_base": False, "low_base_threshold_eur": 5e7,
            },
            "group": {"name": "EV batteries (Li-ion)"},
        })

    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["hs_yoy_imports"]
    headers = [c.value for c in ws[3]]
    row = [[c.value for c in r] for r in ws.iter_rows(min_row=4)][0]
    assert row[headers.index("low_base")] is False
    assert row[headers.index("near_low_base_threshold")] is True


def test_mirror_gaps_includes_cif_fob_baseline_columns(
    empty_findings, test_db_url, tmp_path,
):
    """The mirror_gaps tab now exposes the per-finding CIF/FOB baseline
    + scope + computed excess-over-baseline-pp."""
    with psycopg2.connect(test_db_url) as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO scrape_runs (source_url, status) VALUES "
            "('seed', 'success') RETURNING id"
        )
        run = cur.fetchone()[0]
        detail = {
            "iso2": "NL",
            "gacc": {"partner_label_raw": "Netherlands", "value_eur_converted": 1e10},
            "eurostat": {"total_eur": 1.7e10},
            "gap_eur": 7e9, "gap_pct": 0.65, "is_aggregate": False,
            "cif_fob_baseline": {
                "baseline_pct": 0.0655,
                "scope": "per-partner",
                "partner_iso2": "NL",
                "source": "OECD ITIC dataset 2022 (NL)",
            },
        }
        cur.execute(
            "INSERT INTO findings (scrape_run_id, kind, subkind, observation_ids, "
            "                       score, title, body, detail) "
            "VALUES (%s, 'anomaly', 'mirror_gap', '{}', 0.65, 'NL gap', 'b', %s::jsonb)",
            (run, json.dumps(detail)),
        )
        conn.commit()

    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["mirror_gaps"]
    headers = [c.value for c in ws[3]]
    row = [[c.value for c in r] for r in ws.iter_rows(min_row=4)][0]
    assert row[headers.index("cif_fob_baseline_pct")] == 0.0655
    assert row[headers.index("cif_fob_baseline_scope")] == "per-partner"
    # excess = (|0.65| - 0.0655) * 100 = 58.45 pp
    assert abs(row[headers.index("excess_over_baseline_pp")] - 58.45) < 0.01


def test_link_column_empty_when_permalink_base_unset(empty_findings, test_db_url, tmp_path):
    with psycopg2.connect(test_db_url) as conn:
        _seed_one_finding(conn, "hs_group_yoy", "EV batteries (Li-ion)", {
            "windows": {"current_end": "2026-02-01", "current_start": "2026-02-01"},
            "totals": {"yoy_pct": 0.4, "current_12mo_eur": 1e9, "prior_12mo_eur": 0.9e9,
                       "low_base": False},
            "group": {"name": "EV batteries (Li-ion)"},
        })

    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["hs_yoy_imports"]
    headers = [c.value for c in ws[3]]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=4)]
    assert all(r[headers.index("link")] in (None, "") for r in rows)


def test_link_column_emits_hyperlink_formula_when_permalink_base_set(
    empty_findings, test_db_url, tmp_path, monkeypatch,
):
    monkeypatch.setenv(sheets_export.PERMALINK_BASE_ENV, "https://gacc.example")
    with psycopg2.connect(test_db_url) as conn:
        fid = _seed_one_finding(conn, "hs_group_yoy", "EV batteries (Li-ion)", {
            "windows": {"current_end": "2026-02-01", "current_start": "2026-02-01"},
            "totals": {"yoy_pct": 0.4, "current_12mo_eur": 1e9, "prior_12mo_eur": 0.9e9,
                       "low_base": False},
            "group": {"name": "EV batteries (Li-ion)"},
        })

    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["hs_yoy_imports"]
    headers = [c.value for c in ws[3]]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=4)]
    link = rows[0][headers.index("link")]
    assert link is not None
    assert link.startswith("=HYPERLINK(")
    assert f"finding/{fid}" in link


def test_low_base_review_sheet_only_includes_flagged(empty_findings, test_db_url, tmp_path):
    """The low_base_review sheet must only include findings with low_base=true,
    and must expose scope + flow columns for filtering."""
    with psycopg2.connect(test_db_url) as conn:
        # One flagged (UK scope, export flow), one not flagged (EU-27 imports)
        _seed_one_finding(conn, "hs_group_yoy_uk_export", "Rare-earth materials", {
            "windows": {"current_end": "2026-02-01", "current_start": "2026-02-01"},
            "totals": {"yoy_pct": 0.5, "current_12mo_eur": 1e7, "prior_12mo_eur": 0.5e7,
                       "low_base": True, "low_base_threshold_eur": 5e7},
            "group": {"name": "Rare-earth materials"},
        })
        _seed_one_finding(conn, "hs_group_yoy", "EV batteries (Li-ion)", {
            "windows": {"current_end": "2026-02-01", "current_start": "2026-02-01"},
            "totals": {"yoy_pct": 0.3, "current_12mo_eur": 2e10, "prior_12mo_eur": 1.5e10,
                       "low_base": False},
            "group": {"name": "EV batteries (Li-ion)"},
        })

    out = sheets_export.export(out_format="xlsx", out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["low_base_review"]
    headers = [c.value for c in ws[3]]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=4)]
    groups = [r[headers.index("group")] for r in rows]
    assert "Rare-earth materials" in groups
    assert "EV batteries (Li-ion)" not in groups
    # Scope + flow columns are populated for the flagged row
    flagged = next(r for r in rows if r[headers.index("group")] == "Rare-earth materials")
    assert flagged[headers.index("scope")] == "uk"
    assert flagged[headers.index("flow")] == "export"


def test_decode_subkind_helper():
    from sheets_export import _decode_subkind as fn
    assert fn("hs_group_yoy") == ("eu_27", "import")
    assert fn("hs_group_yoy_export") == ("eu_27", "export")
    assert fn("hs_group_yoy_uk") == ("uk", "import")
    assert fn("hs_group_yoy_uk_export") == ("uk", "export")
    assert fn("hs_group_yoy_combined") == ("eu_27_plus_uk", "import")
    assert fn("hs_group_yoy_combined_export") == ("eu_27_plus_uk", "export")


def test_briefing_pack_export_also_writes_data_xlsx(
    test_db_url, tmp_path, monkeypatch,
):
    """When briefing_pack.export() runs, it should drop a data.xlsx into
    the same per-export folder as findings.md + leads.md so all three
    artefacts share a single DB snapshot."""
    monkeypatch.setenv("DATABASE_URL", test_db_url)
    with psycopg2.connect(test_db_url) as conn, conn.cursor() as cur:
        cur.execute(
            "TRUNCATE findings, observations, source_snapshots, eurostat_raw_rows, "
            "scrape_runs, releases, brief_runs RESTART IDENTITY CASCADE"
        )
    import briefing_pack
    brief_path, leads_path = briefing_pack.export(
        out_dir=str(tmp_path / "20260510-1200"),
    )
    folder = Path(brief_path).parent
    assert (folder / "findings.md").exists()
    assert (folder / "leads.md").exists()
    assert (folder / "data.xlsx").exists()


def test_briefing_pack_export_can_disable_spreadsheet(
    test_db_url, tmp_path, monkeypatch,
):
    """`spreadsheet=False` skips xlsx generation — useful for tests or
    legacy explicit-path callers that don't need it."""
    monkeypatch.setenv("DATABASE_URL", test_db_url)
    with psycopg2.connect(test_db_url) as conn, conn.cursor() as cur:
        cur.execute(
            "TRUNCATE findings, observations, source_snapshots, eurostat_raw_rows, "
            "scrape_runs, releases, brief_runs RESTART IDENTITY CASCADE"
        )
    import briefing_pack
    brief_path, leads_path = briefing_pack.export(
        out_dir=str(tmp_path / "20260510-1200"),
        spreadsheet=False,
    )
    folder = Path(brief_path).parent
    assert (folder / "data.xlsx").exists() is False


def test_google_sheets_writer_raises_until_wired_up():
    with pytest.raises(NotImplementedError, match="not yet wired up"):
        sheets_export.GoogleSheetsWriter().write([], "any-spreadsheet-id")
