"""Spreadsheet export for findings.

Renders findings into a list of SheetData (header + rows) — pure SQL → in-
memory tables — then writes via a pluggable writer. v1 ships an XlsxWriter
that produces local .xlsx files via openpyxl, and a GoogleSheetsWriter stub
that will pick up either a service-account JSON (production) or OAuth user
creds (interactive testing) once we wire it up.

The shape is identical between the two destinations: switching from XLSX to
Sheets is a one-line change of writer, no rebuild of the data layer.

By default, `briefing_pack.export()` invokes this module to drop a
`data.xlsx` into the per-export folder alongside `findings.md` + `leads.md` —
all three artefacts share a single DB snapshot. Standalone CLI usage
(`scrape.py --export-sheet`) is still supported for spreadsheet-only runs.

Permalink convention: every output sheet has a `finding_id` column (the
canonical reference) and a `link` column. The link column emits a Sheets
HYPERLINK formula based on the GACC_PERMALINK_BASE env var. If unset, the
column stays empty — no rotten links. When a web UI exists later, set
GACC_PERMALINK_BASE and the column lights up automatically (formula resolves
at view time, not at export time).

Editorial design: this spreadsheet is intentionally LLM-free. The
narrative_hs_group findings (LLM-scaffolded leads) live in the
companion `leads.md` document, NOT in any tab here. A downstream
data-journalism workflow filtering / pivoting / charting from the
spreadsheet should be reasoning over deterministic findings, not over
another LLM's interpretation of them.
"""

import logging
import os
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Protocol

import psycopg2
import psycopg2.extras

from briefing_pack import (
    _ALL_UNIVERSAL_CAVEATS,
    _SCOPE_LABEL,
    _SCOPE_SUBKIND_SUFFIX,
    _compute_predictability_per_group,
    _compute_top_movers,
    is_threshold_fragile,
)

log = logging.getLogger(__name__)

PERMALINK_BASE_ENV = "GACC_PERMALINK_BASE"

SCOPES = ("eu_27", "uk", "eu_27_plus_uk")


@dataclass
class SheetData:
    """One sheet's worth of data, ready to render."""
    name: str            # sheet/tab name, max 31 chars for Excel
    description: str     # one-line description rendered as the first row
    headers: list[str]
    rows: list[list[Any]]


class Writer(Protocol):
    def write(self, sheets: list[SheetData], dest: str) -> str:
        """Write sheets to dest. Returns the final path or URL."""
        ...


def _conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def _link_cell(finding_id: int) -> str:
    """Return a HYPERLINK formula if GACC_PERMALINK_BASE is set, else empty
    string. The formula resolves at view-time, so existing exports light up
    once the base is configured later."""
    base = os.environ.get(PERMALINK_BASE_ENV, "").rstrip("/")
    if not base:
        return ""
    return f'=HYPERLINK("{base}/finding/{finding_id}", "open")'


def _yoy_subkind(scope: str, flow: int) -> str:
    """Compose hs_group_yoy subkind from (scope, flow). Mirrors the
    convention in briefing_pack._section_hs_yoy_movers."""
    return f"hs_group_yoy{_SCOPE_SUBKIND_SUFFIX[scope]}{'' if flow == 1 else '_export'}"


def _trajectory_subkind(scope: str, flow: int) -> str:
    return f"hs_group_trajectory{_SCOPE_SUBKIND_SUFFIX[scope]}{'' if flow == 1 else '_export'}"


def _filter_visible_caveats(codes: list[str] | None) -> list[str]:
    """Drop the family-universal caveats (defined once in the brief
    methodology footer); keep the per-finding-variable ones."""
    return [c for c in (codes or []) if c not in _ALL_UNIVERSAL_CAVEATS]


# =============================================================================
# Sheet builders — one function per output tab
# =============================================================================


def assemble_sheets() -> list[SheetData]:
    """Build the journalist-facing spreadsheet tabs.

    Tab roster (10):

    1. summary — wide, one row per HS group, all 3 scopes × 2 flows side
       by side. Best starting place for editorial scanning.
    2. hs_yoy_imports — long, one row per (group, scope), full metric
       set with predictability + fragility annotations.
    3. hs_yoy_exports — same shape, flow=2.
    4. hs_yoy_reporter_movers — long, one row per (group, scope, flow,
       reporter); contribution + share of group's delta. Phase 6.11.
    5. trajectories — long, one row per (group, scope, flow), shape
       classification + features.
    6. gacc_bilateral_yoy — one row per (partner, flow) of the GACC-side
       bilateral aggregate findings (Tier 2 of findings.md in sortable
       form). Carries visible_caveats + jan_feb_combined_years so
       Chinese-New-Year combined-release coverage is filterable.
    7. mirror_gaps — latest mirror_gap per partner, with per-country
       CIF/FOB baseline + excess-over-baseline split.
    8. mirror_gap_movers — z-score sheet, sorted by |z|.
    9. low_base_review — findings flagged low_base; pre-quote audit queue.
    10. predictability_index — per-group YoY predictability (Phase 6.6
        backtest output) summarised so a journalist can sort/filter on
        which groups give robust headline percentages.

    The narrative_hs_group findings (LLM-scaffolded leads) are
    intentionally excluded — they live in `leads.md` alongside this
    spreadsheet in the same export folder."""
    # Compute predictability + top-5 movers once and pass into the
    # YoY tabs that surface them.
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        predictability = _compute_predictability_per_group(cur)
        top_movers_rows = _compute_top_movers(cur, predictability=predictability)
    # Map finding_id → (rank, score) for direct row lookup in the long
    # sheets. The summary sheet needs (group_name, flow) → rank lookup
    # since it's one-row-per-group; build that separately.
    top_movers_by_id: dict[int, tuple[int, float]] = {
        m["id"]: (rank, m["score"])
        for rank, m in enumerate(top_movers_rows, start=1)
    }
    top_movers_by_group_flow: dict[tuple[str, int], int] = {}
    for rank, m in enumerate(top_movers_rows, start=1):
        flow = 2 if m["subkind"].endswith("_export") else 1
        top_movers_by_group_flow[(m["group_name"], flow)] = rank

    sheets: list[SheetData] = []
    sheets.append(_summary_sheet(predictability, top_movers_by_group_flow))
    sheets.append(_hs_yoy_long_sheet(flow=1, predictability=predictability,
                                      top_movers_by_id=top_movers_by_id))
    sheets.append(_hs_yoy_long_sheet(flow=2, predictability=predictability,
                                      top_movers_by_id=top_movers_by_id))
    sheets.append(_hs_yoy_reporter_movers_sheet())
    sheets.append(_trajectories_long_sheet())
    sheets.append(_gacc_bilateral_yoy_sheet())
    sheets.append(_mirror_gaps_sheet())
    sheets.append(_mirror_gap_movers_sheet())
    sheets.append(_low_base_review_sheet())
    sheets.append(_predictability_index_sheet(predictability))
    return sheets


def _summary_sheet(
    predictability: dict[str, tuple[str, float, int]],
    top_movers_by_group_flow: dict[tuple[str, int], int] | None = None,
) -> SheetData:
    """One row per HS group; all three scopes × both flows side by side
    so a journalist can compare EU-27 / UK / combined views at a glance.
    Trajectory shape (EU-27, both flows) anchors the row's narrative
    classification.

    Wide-format trade-off: lots of columns, no per-row finding_ids
    (those live in the long-format tabs below). Use this tab for
    scanning; drill to the long tabs for detail.

    `top_movers_by_group_flow` (optional): (group_name, flow) → rank
    for this cycle's editorial picks. Surfaced as two right-most
    columns — `top_movers_rank_imp` and `top_movers_rank_exp`. NULL
    when the group's import / export flow didn't make the picks.
    """
    top_movers_by_group_flow = top_movers_by_group_flow or {}
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        # Build one CTE per (scope, flow) → 6 CTEs. Each picks the latest
        # finding per group. Then LEFT JOIN them all onto hs_groups so the
        # row exists even if a group has no findings in some scope.
        ctes = []
        select_cols = []
        for scope in SCOPES:
            for flow in (1, 2):
                sk = _yoy_subkind(scope, flow)
                alias = f"{scope}_{'imp' if flow == 1 else 'exp'}"
                ctes.append(f"""
                    {alias} AS (
                      SELECT DISTINCT ON (detail->'group'->>'name')
                             detail->'group'->>'name' AS group_name,
                             (detail->'totals'->>'current_12mo_eur')::numeric AS eur,
                             (detail->'totals'->>'yoy_pct')::numeric AS yoy,
                             (detail->'totals'->>'low_base')::boolean AS low_base
                        FROM findings
                       WHERE subkind = '{sk}' AND superseded_at IS NULL
                    ORDER BY detail->'group'->>'name', (detail->'windows'->>'current_end')::date DESC, id DESC
                    )
                """)
                select_cols.extend([
                    f"{alias}.eur AS {alias}_eur",
                    f"{alias}.yoy AS {alias}_yoy",
                    f"{alias}.low_base AS {alias}_lb",
                ])
        # Trajectory shapes for the EU-27 view (a single column per flow).
        ctes.append("""
            traj_imp AS (
              SELECT DISTINCT ON (detail->'group'->>'name')
                     detail->'group'->>'name' AS group_name,
                     detail->>'shape' AS shape
                FROM findings
               WHERE subkind = 'hs_group_trajectory' AND superseded_at IS NULL
            ORDER BY detail->'group'->>'name', created_at DESC
            )
        """)
        ctes.append("""
            traj_exp AS (
              SELECT DISTINCT ON (detail->'group'->>'name')
                     detail->'group'->>'name' AS group_name,
                     detail->>'shape' AS shape
                FROM findings
               WHERE subkind = 'hs_group_trajectory_export' AND superseded_at IS NULL
            ORDER BY detail->'group'->>'name', created_at DESC
            )
        """)
        joins = []
        for scope in SCOPES:
            for flow in (1, 2):
                alias = f"{scope}_{'imp' if flow == 1 else 'exp'}"
                joins.append(f"LEFT JOIN {alias} ON {alias}.group_name = g.name")
        joins.append("LEFT JOIN traj_imp ON traj_imp.group_name = g.name")
        joins.append("LEFT JOIN traj_exp ON traj_exp.group_name = g.name")

        sql = (
            "WITH " + ",".join(ctes) + "\n"
            "SELECT g.name AS group_name, " + ", ".join(select_cols) +
            ", traj_imp.shape AS traj_imp_shape, traj_exp.shape AS traj_exp_shape\n"
            "  FROM hs_groups g\n"
            + "\n".join(joins) + "\n"
            "ORDER BY g.id"
        )
        cur.execute(sql)
        raw_rows = cur.fetchall()

    headers = ["group", "predictability_badge", "predictability_pct"]
    for scope in SCOPES:
        scope_short = {"eu_27": "eu27", "uk": "uk", "eu_27_plus_uk": "combined"}[scope]
        for flow_short in ("imp", "exp"):
            headers.extend([
                f"{scope_short}_{flow_short}_yoy_pct",
                f"{scope_short}_{flow_short}_eur_12mo",
                f"{scope_short}_{flow_short}_low_base",
            ])
    headers.extend([
        "eu27_imp_trajectory_shape", "eu27_exp_trajectory_shape",
        "top_movers_rank_imp", "top_movers_rank_exp",
    ])

    rows: list[list[Any]] = []
    for r in raw_rows:
        gn = r["group_name"]
        pred = predictability.get(gn)
        badge = pred[0] if pred else ""
        pred_pct = round(pred[1] * 100, 0) if pred else None
        row: list[Any] = [gn, badge, pred_pct]
        for scope in SCOPES:
            for flow_short in ("imp", "exp"):
                alias = f"{scope}_{flow_short}"
                row.extend([
                    _to_float(r[f"{alias}_yoy"]),
                    _to_float(r[f"{alias}_eur"]),
                    bool(r[f"{alias}_lb"]) if r[f"{alias}_lb"] is not None else None,
                ])
        row.extend([
            r["traj_imp_shape"], r["traj_exp_shape"],
            top_movers_by_group_flow.get((gn, 1)),  # rank for imports
            top_movers_by_group_flow.get((gn, 2)),  # rank for exports
        ])
        rows.append(row)

    return SheetData(
        name="summary",
        description=(
            "One row per HS group; all three comparison scopes (EU-27 / UK / "
            "combined) and both flows (imports / exports) side by side. "
            "Predictability badge: 🟢 = headline % is robust over 6mo, "
            "🟡 = noisy, 🔴 = volatile (lean on trajectory shape instead). "
            "`top_movers_rank_imp` / `top_movers_rank_exp` are 1-5 for "
            "this cycle's editorial picks — the same rows that lead "
            "`findings.md`; NULL otherwise. Companion documents in this "
            "folder: findings.md (deterministic Markdown rendering of the "
            "same findings, NotebookLM-ready) and leads.md (LLM-scaffolded "
            "investigation starts per HS group)."
        ),
        headers=headers, rows=rows,
    )


def _hs_yoy_long_sheet(
    flow: int, predictability: dict[str, tuple[str, float, int]],
    top_movers_by_id: dict[int, tuple[int, float]] | None = None,
) -> SheetData:
    """Long format: one row per (group, scope), latest period only.
    Three scopes stacked → ~3x rows of the old EU-27-only sheet, with
    `scope` as a filterable column. Adds predictability badge +
    threshold-fragility flag + top-movers rank/score per row.

    `top_movers_by_id` (optional): finding_id → (rank, score) for this
    cycle's Top-5 picks (see briefing_pack._compute_top_movers).
    Surfaced as `top_movers_rank` (1-5 or empty) and `top_movers_score`
    (composite numeric, also populated for rows that didn't quite make
    the cut so the journalist can see the long tail).
    """
    top_movers_by_id = top_movers_by_id or {}
    flow_label = "imports (CN→reporter)" if flow == 1 else "exports (reporter→CN)"
    name = f"hs_yoy_{'imports' if flow == 1 else 'exports'}"
    raw_rows: list[tuple] = []
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        for scope in SCOPES:
            sk = _yoy_subkind(scope, flow)
            cur.execute(
                """
                SELECT DISTINCT ON (detail->'group'->>'name')
                    id, score,
                    detail->'group'->>'name' AS group_name,
                    (detail->'windows'->>'current_end')::date AS period,
                    (detail->'totals'->>'current_12mo_eur')::numeric AS current_eur,
                    (detail->'totals'->>'prior_12mo_eur')::numeric AS prior_eur,
                    (detail->'totals'->>'yoy_pct')::numeric AS yoy_pct,
                    (detail->'totals'->>'current_12mo_kg')::numeric AS current_kg,
                    (detail->'totals'->>'yoy_pct_kg')::numeric AS yoy_kg_pct,
                    (detail->'totals'->>'current_unit_price_eur_per_kg')::numeric AS unit_price,
                    (detail->'totals'->>'unit_price_pct_change')::numeric AS unit_price_pct,
                    (detail->'totals'->>'low_base')::boolean AS low_base,
                    (detail->'totals'->>'low_base_threshold_eur')::numeric AS lb_threshold,
                    (detail->'totals'->>'partial_window')::boolean AS partial_window,
                    (detail->'totals'->>'kg_coverage_pct')::numeric AS kg_coverage
                  FROM findings
                 WHERE subkind = %s AND superseded_at IS NULL
              ORDER BY detail->'group'->>'name', (detail->'windows'->>'current_end')::date DESC, id DESC
                """,
                (sk,),
            )
            for r in cur.fetchall():
                raw_rows.append((scope, r))

    import math
    headers = [
        "finding_id", "link", "group", "scope", "period",
        "current_12mo_eur", "prior_12mo_eur", "yoy_pct",
        "current_12mo_kg", "yoy_kg_pct",
        "unit_price_eur_per_kg", "unit_price_pct",
        "low_base", "near_low_base_threshold",
        "predictability_badge", "predictability_pct",
        "kg_coverage_pct", "partial_window",
        "top_movers_rank", "top_movers_score",
        "score",
    ]
    rows = []
    for scope, r in raw_rows:
        pred = predictability.get(r["group_name"])
        badge = pred[0] if pred else ""
        pred_pct = round(pred[1] * 100, 0) if pred else None
        # Top-movers rank: present only for the 5 picks (eu_27 scope,
        # both flows, filtered + composite-ranked — see
        # briefing_pack._compute_top_movers).
        rank_score = top_movers_by_id.get(r["id"])
        top_rank = rank_score[0] if rank_score else None
        # Composite score for every row, even non-picks — lets the
        # journalist sort by score and see the long tail under the
        # editorial threshold. NULL when yoy or eur is missing.
        yoy = r["yoy_pct"]
        eur = r["current_eur"]
        top_score: float | None
        if yoy is not None and eur is not None and float(eur) > 0:
            top_score = abs(float(yoy)) * math.log10(max(float(eur), 1.0))
        else:
            top_score = None
        rows.append([
            r["id"], _link_cell(r["id"]), r["group_name"], scope, r["period"].isoformat(),
            _to_float(r["current_eur"]), _to_float(r["prior_eur"]), _to_float(r["yoy_pct"]),
            _to_float(r["current_kg"]), _to_float(r["yoy_kg_pct"]),
            _to_float(r["unit_price"]), _to_float(r["unit_price_pct"]),
            bool(r["low_base"]) if r["low_base"] is not None else False,
            is_threshold_fragile(r["current_eur"], r["prior_eur"], r["lb_threshold"]),
            badge, pred_pct,
            _to_float(r["kg_coverage"]),
            bool(r["partial_window"]) if r["partial_window"] is not None else False,
            top_rank,
            round(top_score, 4) if top_score is not None else None,
            _to_float(r["score"]),
        ])
    return SheetData(
        name=name,
        description=(
            f"Latest rolling-12mo {flow_label} per HS group, all scopes "
            "(EU-27 / UK / EU-27 + UK combined). Filter on `scope` to "
            "narrow. `near_low_base_threshold` = TRUE means the finding "
            "is within 1.5x the low_base threshold; classification is "
            "fragile to small threshold changes. `top_movers_rank` is "
            "populated (1-5) for the cycle's editorial picks — the same "
            "rows that lead `findings.md`. `top_movers_score` "
            "(|yoy_pct| × log10(current_eur)) is computed on every row "
            "so you can sort by it and see the long tail."
        ),
        headers=headers, rows=rows,
    )


def _hs_yoy_reporter_movers_sheet() -> SheetData:
    """Long format: one row per (group, scope, flow, reporter, period).

    Phase 6.11. Drawn from `detail.per_reporter_breakdown` on each
    `hs_group_yoy*` finding. The brief surfaces the top 5 reporters per
    mover; this tab surfaces the top 10 across every group / scope / flow,
    making it the right place for "which member state moved the needle on
    group X" filter-and-sort queries (Soapbox A5.6 / A4.5: "Germany alone
    accounts for 66% of the EU-wide drop").

    Each row carries the originating finding_id so the spreadsheet pairs
    one-to-many with the `hs_yoy_imports` / `hs_yoy_exports` tabs."""
    raw: list[tuple[str, int, dict]] = []
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        for scope in SCOPES:
            for flow in (1, 2):
                sk = _yoy_subkind(scope, flow)
                cur.execute(
                    """
                    SELECT DISTINCT ON (detail->'group'->>'name')
                        id,
                        detail->'group'->>'name' AS group_name,
                        (detail->'windows'->>'current_end')::date AS period,
                        detail->'per_reporter_breakdown' AS breakdown
                      FROM findings
                     WHERE subkind = %s AND superseded_at IS NULL
                  ORDER BY detail->'group'->>'name', (detail->'windows'->>'current_end')::date DESC, id DESC
                    """,
                    (sk,),
                )
                for r in cur.fetchall():
                    raw.append((scope, flow, dict(r)))

    headers = [
        "finding_id", "link", "group", "scope", "flow", "period",
        "reporter", "rank_by_abs_delta",
        "current_12mo_eur", "prior_12mo_eur", "delta_eur",
        "yoy_pct", "yoy_pct_kg", "share_of_group_delta_pct",
        "current_12mo_kg", "prior_12mo_kg",
    ]
    rows: list[list[Any]] = []
    for scope, flow, r in raw:
        breakdown = r["breakdown"] or []
        period_str = r["period"].isoformat() if r["period"] else None
        for rank, pr in enumerate(breakdown, start=1):
            rows.append([
                r["id"], _link_cell(r["id"]), r["group_name"], scope,
                "import" if flow == 1 else "export", period_str,
                pr.get("reporter"), rank,
                _to_float(pr.get("current_eur")),
                _to_float(pr.get("prior_eur")),
                _to_float(pr.get("delta_eur")),
                _to_float(pr.get("yoy_pct")),
                _to_float(pr.get("yoy_pct_kg")),
                _to_float(pr.get("share_of_group_delta_pct")),
                _to_float(pr.get("current_kg")),
                _to_float(pr.get("prior_kg")),
            ])
    return SheetData(
        name="hs_yoy_reporter_movers",
        description=(
            "Per-reporter contributions to each HS group's 12mo YoY. "
            "One row per (group, scope, flow, reporter); up to 10 reporters "
            "per finding, ranked by absolute EUR delta. `share_of_group_delta_pct` "
            "is the reporter's contribution to the group's overall delta — "
            "positive = pushed in the group's direction, negative = pushed "
            "against. Filter `scope` / `flow` to narrow."
        ),
        headers=headers, rows=rows,
    )


def _trajectories_long_sheet() -> SheetData:
    """Long format: one row per (group, scope, flow). Shape + features."""
    raw_rows: list[tuple] = []
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        for scope in SCOPES:
            for flow in (1, 2):
                sk = _trajectory_subkind(scope, flow)
                cur.execute(
                    """
                    SELECT DISTINCT ON (detail->'group'->>'name')
                        id, score,
                        detail->'group'->>'name' AS group_name,
                        detail->>'shape' AS shape,
                        detail->>'shape_label' AS shape_label,
                        (detail->'features'->>'last_yoy')::numeric AS last_yoy,
                        (detail->'features'->>'max_yoy')::numeric AS peak,
                        (detail->'features'->>'min_yoy')::numeric AS trough,
                        (detail->'features'->>'low_base_majority')::boolean AS low_base_majority,
                        (detail->'features'->>'has_strong_seasonal_signal')::boolean AS seasonal,
                        (detail->'features'->>'effective_series_length')::int AS n_windows
                      FROM findings
                     WHERE subkind = %s AND superseded_at IS NULL
                  ORDER BY detail->'group'->>'name', created_at DESC
                    """,
                    (sk,),
                )
                for r in cur.fetchall():
                    raw_rows.append((scope, flow, r))

    headers = [
        "finding_id", "link", "group", "scope", "flow",
        "shape", "shape_label", "latest_yoy_pct", "peak_yoy_pct", "trough_yoy_pct",
        "n_windows", "low_base_majority", "seasonal_signal",
    ]
    rows = []
    for scope, flow, r in raw_rows:
        rows.append([
            r["id"], _link_cell(r["id"]), r["group_name"], scope,
            "import" if flow == 1 else "export",
            r["shape"], r["shape_label"],
            _to_float(r["last_yoy"]), _to_float(r["peak"]), _to_float(r["trough"]),
            r["n_windows"],
            bool(r["low_base_majority"]) if r["low_base_majority"] is not None else None,
            bool(r["seasonal"]) if r["seasonal"] is not None else False,
        ])
    return SheetData(
        name="trajectories",
        description=(
            "Trajectory shape per (HS group, scope, flow). Shape vocabulary: "
            "rising / falling (+ accel/decel), u_recovery, inverse_u_peak, "
            "dip_recovery, failed_recovery, volatile, flat. Filter on `scope` "
            "or `flow` to narrow. `seasonal_signal` = TRUE means the series "
            "has a strong lag-12 autocorrelation (deseasonalise before "
            "interpreting raw movement)."
        ),
        headers=headers, rows=rows,
    )


def _mirror_gaps_sheet() -> SheetData:
    """Latest mirror_gap per partner, with per-country CIF/FOB baseline
    expansion. Adds: baseline_pct, baseline_scope (per-partner /
    global), excess_over_baseline_pp, transshipment_hub flag, visible
    caveats (universal codes suppressed; per-finding ones kept)."""
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute(
            """
            SELECT DISTINCT ON (detail->>'iso2')
                f.id, f.score,
                detail->>'iso2' AS iso2,
                detail->'gacc'->>'partner_label_raw' AS gacc_label,
                (detail->'gacc'->>'value_eur_converted')::numeric AS gacc_eur,
                (detail->'eurostat'->>'total_eur')::numeric AS eurostat_eur,
                (detail->>'gap_eur')::numeric AS gap_eur,
                (detail->>'gap_pct')::numeric AS gap_pct,
                (detail->>'is_aggregate')::boolean AS is_aggregate,
                (detail->'cif_fob_baseline'->>'baseline_pct')::numeric AS baseline_pct,
                detail->'cif_fob_baseline'->>'scope' AS baseline_scope,
                detail->'cif_fob_baseline'->>'source' AS baseline_source,
                detail->'transshipment_hub'->>'iso2' AS hub_iso2,
                detail->'caveat_codes' AS caveat_codes,
                (SELECT to_char(r.period, 'YYYY-MM')
                   FROM observations o JOIN releases r ON r.id = o.release_id
                  WHERE o.id = f.observation_ids[1]) AS period
              FROM findings f
             WHERE subkind = 'mirror_gap' AND superseded_at IS NULL
          ORDER BY detail->>'iso2',
                   (SELECT r.period FROM observations o JOIN releases r ON r.id = o.release_id
                     WHERE o.id = f.observation_ids[1]) DESC,
                   id DESC
            """
        )
        rows_raw = cur.fetchall()

    headers = [
        "finding_id", "link", "iso2", "gacc_label", "period",
        "gacc_eur_converted", "eurostat_eur", "gap_eur", "gap_pct",
        "cif_fob_baseline_pct", "cif_fob_baseline_scope",
        "excess_over_baseline_pp", "is_transshipment_hub",
        "is_aggregate", "visible_caveats", "score",
    ]
    rows = []
    for r in rows_raw:
        gap_pct = float(r["gap_pct"]) if r["gap_pct"] is not None else None
        baseline_pct = float(r["baseline_pct"]) if r["baseline_pct"] is not None else None
        excess_pp = (
            (abs(gap_pct) - baseline_pct) * 100
            if gap_pct is not None and baseline_pct is not None else None
        )
        visible = _filter_visible_caveats(r["caveat_codes"] or [])
        rows.append([
            r["id"], _link_cell(r["id"]), r["iso2"], r["gacc_label"], r["period"],
            _to_float(r["gacc_eur"]), _to_float(r["eurostat_eur"]),
            _to_float(r["gap_eur"]), gap_pct,
            baseline_pct, r["baseline_scope"],
            round(excess_pp, 2) if excess_pp is not None else None,
            r["hub_iso2"] is not None,
            bool(r["is_aggregate"]) if r["is_aggregate"] is not None else False,
            ", ".join(visible),
            _to_float(r["score"]),
        ])
    return SheetData(
        name="mirror_gaps",
        description=(
            "China-export-to-X (GACC, EUR-converted) vs X-import-from-China "
            "(Eurostat). Per-country CIF/FOB baseline from OECD ITIC; "
            "excess_over_baseline_pp = (|gap_pct| - baseline_pct) in "
            "percentage points (the part that isn't structural freight + "
            "insurance). is_transshipment_hub = TRUE for known routing "
            "hubs (NL, BE, HK, SG, AE, MX) where gap mostly reflects "
            "transit, not direct trade."
        ),
        headers=headers, rows=rows,
    )


def _mirror_gap_movers_sheet() -> SheetData:
    """Mirror-gap z-score findings sorted by |z|. Same as before but
    with visible caveats column."""
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute(
            """
            SELECT id, score,
                   detail->>'iso2' AS iso2,
                   to_char((detail->>'period')::date, 'YYYY-MM') AS period,
                   (detail->>'gap_pct')::numeric AS gap_pct,
                   (detail->'baseline'->>'mean')::numeric AS baseline_mean,
                   (detail->'baseline'->>'stdev')::numeric AS baseline_stdev,
                   (detail->'baseline'->>'n')::int AS baseline_n,
                   (detail->>'z_score')::numeric AS z,
                   detail->'caveat_codes' AS caveat_codes
              FROM findings
             WHERE subkind = 'mirror_gap_zscore' AND superseded_at IS NULL
          ORDER BY abs((detail->>'z_score')::numeric) DESC
             LIMIT 50
            """
        )
        rows_raw = cur.fetchall()

    headers = [
        "finding_id", "link", "iso2", "period", "gap_pct",
        "baseline_mean_pct", "baseline_stdev_pct", "baseline_n", "z_score",
        "visible_caveats",
    ]
    rows = []
    for r in rows_raw:
        visible = _filter_visible_caveats(r["caveat_codes"] or [])
        rows.append([
            r["id"], _link_cell(r["id"]), r["iso2"], r["period"],
            _to_float(r["gap_pct"]),
            _to_float(r["baseline_mean"]), _to_float(r["baseline_stdev"]),
            r["baseline_n"], _to_float(r["z"]),
            ", ".join(visible),
        ])
    return SheetData(
        name="mirror_gap_movers",
        description=(
            "Mirror-trade gap shifts vs each partner's own rolling baseline. "
            "Sorted by |z|. High |z| = gap moved unusually for that partner. "
            "z_score >= 1.5 was the analyser threshold; |z| >= 2.5 is "
            "robust, [1.5, 2.0) is marginal — see methodology.md §7."
        ),
        headers=headers, rows=rows,
    )


def _gacc_bilateral_yoy_sheet() -> SheetData:
    """One row per (partner, flow) for the latest gacc_bilateral_aggregate_yoy
    finding. Carries the same three YoY operators the brief's Tier 2 block
    surfaces (12mo rolling / YTD / single-month) plus a `visible_caveats`
    column and an explicit `jan_feb_combined_years` column.

    The `jan_feb_combined` caveat tells the journalist that part of the
    12mo total came in as a 2-month cumulative chunk (GACC's Chinese-New-
    Year combined release), not as separate monthly figures. Filtering or
    sorting the spreadsheet on that column lets a journalist isolate the
    rows that rest on the cumulative — useful when deciding whether to
    quote a per-month series claim alongside the 12mo headline."""
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute(
            """
            SELECT DISTINCT ON (detail->'partner'->>'raw_label', subkind)
                   id,
                   detail->'partner'->>'raw_label' AS partner_label,
                   detail->'partner'->>'kind' AS partner_kind,
                   subkind,
                   (detail->'windows'->>'current_end')::date AS current_end,
                   (detail->'totals'->>'current_12mo_eur')::numeric AS rolling_curr_eur,
                   (detail->'totals'->>'prior_12mo_eur')::numeric AS rolling_prior_eur,
                   (detail->'totals'->>'yoy_pct')::numeric AS rolling_yoy_pct,
                   (detail->'totals'->'ytd_cumulative'->>'yoy_pct')::numeric AS ytd_yoy_pct,
                   (detail->'totals'->'ytd_cumulative'->>'current_eur')::numeric AS ytd_curr_eur,
                   (detail->'totals'->'ytd_cumulative'->>'months_in_ytd')::int AS ytd_months,
                   (detail->'totals'->'single_month'->>'yoy_pct')::numeric AS sm_yoy_pct,
                   (detail->'totals'->'single_month'->>'current_eur')::numeric AS sm_curr_eur,
                   (detail->'totals'->>'partial_window')::boolean AS partial_window,
                   detail->'totals'->'jan_feb_combined_years' AS jan_feb_combined_years,
                   detail->'caveat_codes' AS caveat_codes
              FROM findings
             WHERE subkind LIKE 'gacc_bilateral_aggregate_yoy%%'
               AND superseded_at IS NULL
          ORDER BY detail->'partner'->>'raw_label', subkind,
                   (detail->'windows'->>'current_end')::date DESC, id DESC
            """
        )
        rows_raw = list(cur.fetchall())

    headers = [
        "finding_id", "link", "partner", "partner_kind", "flow",
        "current_end", "rolling_12mo_eur", "rolling_yoy_pct",
        "ytd_yoy_pct", "ytd_current_eur", "ytd_months",
        "single_month_yoy_pct", "single_month_current_eur",
        "partial_window", "jan_feb_combined_years", "visible_caveats",
    ]
    rows = []
    # Sort EU bloc first, then single countries alphabetically — mirrors
    # the brief's Tier 2 ordering so a journalist switching surfaces sees
    # the same row order.
    rows_raw.sort(
        key=lambda r: (
            0 if r["partner_kind"] == "eu_bloc" else 1,
            r["partner_label"] or "",
            r["subkind"],
        ),
    )
    for r in rows_raw:
        flow = "export" if r["subkind"] == "gacc_bilateral_aggregate_yoy" else "import"
        visible = _filter_visible_caveats(r["caveat_codes"] or [])
        jfc_years = r["jan_feb_combined_years"] or []
        rows.append([
            r["id"], _link_cell(r["id"]),
            r["partner_label"], r["partner_kind"] or "single_country", flow,
            r["current_end"].isoformat() if r["current_end"] else None,
            _to_float(r["rolling_curr_eur"]), _to_float(r["rolling_yoy_pct"]),
            _to_float(r["ytd_yoy_pct"]), _to_float(r["ytd_curr_eur"]),
            r["ytd_months"],
            _to_float(r["sm_yoy_pct"]), _to_float(r["sm_curr_eur"]),
            bool(r["partial_window"]),
            ", ".join(str(y) for y in jfc_years),
            ", ".join(visible),
        ])
    return SheetData(
        name="gacc_bilateral_yoy",
        description=(
            "GACC-side bilateral YoY for the EU bloc plus every single-"
            "country GACC partner. China's-perspective flow direction: "
            "export=China sells to partner, import=China buys from "
            "partner. Three YoY operators side-by-side: 12mo rolling "
            "(the analyser's primary), YTD cumulative (Soapbox / Merics "
            "register: \"China-X trade +Y% Jan-N YoY\"), and single-"
            "month (Soapbox A3 register). visible_caveats and "
            "jan_feb_combined_years flag when a row's 12mo total "
            "includes a Jan+Feb cumulative chunk rather than separate "
            "monthly figures — see the per-finding provenance file for "
            "the full editorial guidance."
        ),
        headers=headers, rows=rows,
    )


def _low_base_review_sheet() -> SheetData:
    """All hs_group_yoy* findings flagged low_base. Pre-quote audit
    queue. Now includes scope so a journalist can filter by EU-27 / UK
    / combined."""
    with _conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute(
            """
            SELECT id, score, subkind,
                   detail->'group'->>'name' AS group_name,
                   to_char((detail->'windows'->>'current_end')::date, 'YYYY-MM') AS period,
                   (detail->'totals'->>'current_12mo_eur')::numeric AS current_eur,
                   (detail->'totals'->>'prior_12mo_eur')::numeric AS prior_eur,
                   (detail->'totals'->>'yoy_pct')::numeric AS yoy_pct,
                   (detail->'totals'->>'low_base_threshold_eur')::numeric AS threshold
              FROM findings
             WHERE subkind LIKE 'hs_group_yoy%%'
               AND (detail->'totals'->>'low_base')::boolean = true
               AND superseded_at IS NULL
          ORDER BY abs((detail->'totals'->>'yoy_pct')::numeric) DESC
            """
        )
        rows_raw = cur.fetchall()

    headers = [
        "finding_id", "link", "subkind", "scope", "flow",
        "group", "period",
        "current_12mo_eur", "prior_12mo_eur", "yoy_pct", "low_base_threshold_eur",
    ]
    rows = []
    for r in rows_raw:
        scope, flow = _decode_subkind(r["subkind"])
        rows.append([
            r["id"], _link_cell(r["id"]), r["subkind"], scope, flow,
            r["group_name"], r["period"],
            _to_float(r["current_eur"]), _to_float(r["prior_eur"]),
            _to_float(r["yoy_pct"]), _to_float(r["threshold"]),
        ])
    return SheetData(
        name="low_base_review",
        description=(
            "Findings flagged low-base. Verify the absolute figures before "
            "quoting any YoY percentage — small denominators can exaggerate. "
            "Filter `scope` and `flow` to focus."
        ),
        headers=headers, rows=rows,
    )


def _predictability_index_sheet(
    predictability: dict[str, tuple[str, float, int]],
) -> SheetData:
    """One row per HS group: predictability badge + persistence rate
    over the last 6 months. Output of the Phase 6.6 backtest, surfaced
    so a data journalist can sort/filter on which groups give robust
    headline percentages vs which need trajectory-shape framing."""
    headers = [
        "group", "predictability_badge", "predictability_label",
        "persistent_pct", "n_permutations",
    ]
    rows = []
    for gn, (badge, pct, n) in sorted(
        predictability.items(),
        key=lambda kv: (kv[1][0] != "🟢", kv[1][0] != "🟡", -kv[1][1]),
    ):
        label = (
            "persistent" if badge == "🟢"
            else "noisy" if badge == "🟡"
            else "volatile"
        )
        rows.append([gn, badge, label, round(pct * 100, 0), n])

    description = (
        "Per-HS-group YoY-signal stability over the most recent 6 months. "
        "Pairs each group's hs_group_yoy* finding at the latest anchor "
        "period (T) against the same (group, subkind) at T-6 and asks: "
        "did the signal age well? 🟢 ≥67% persistent, 🟡 33-67%, 🔴 <33%. "
        "Editorial use: 🔴 groups need trajectory-shape framing, not a "
        "headline percentage. Empty if no T-6 history exists yet (fresh DB)."
    )
    return SheetData(
        name="predictability_index",
        description=description, headers=headers, rows=rows,
    )


def _decode_subkind(subkind: str) -> tuple[str, str]:
    """Map an hs_group_yoy* subkind string to (scope, flow) pair so the
    spreadsheet can expose them as separate columns."""
    if subkind.endswith("_export"):
        flow = "export"
        prefix = subkind[: -len("_export")]
    else:
        flow = "import"
        prefix = subkind
    if prefix.endswith("_uk"):
        scope = "uk"
    elif prefix.endswith("_combined"):
        scope = "eu_27_plus_uk"
    else:
        scope = "eu_27"
    return scope, flow


def _to_float(v: Any) -> float | None:
    """psycopg2 returns NUMERIC as Decimal; coerce to float for spreadsheet output."""
    if v is None:
        return None
    return float(v)


# =============================================================================
# Writers
# =============================================================================


class XlsxWriter:
    """Render SheetData list to a local .xlsx via openpyxl. The format is
    structurally identical to what GoogleSheetsWriter will push, so the
    iteration on data shape during prototyping carries over directly."""

    def write(self, sheets: list[SheetData], dest: str) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        for sd in sheets:
            ws = wb.create_sheet(title=sd.name[:31])
            ws.append([sd.description])
            ws["A1"].font = Font(italic=True)
            ws.append([])
            ws.append(sd.headers)
            for cell in ws[3]:
                cell.font = Font(bold=True)
            for row in sd.rows:
                ws.append(row)
            # Roughly auto-size columns (cap at 60 chars)
            for col_idx, header in enumerate(sd.headers, start=1):
                max_len = max(
                    [len(str(row[col_idx - 1])) for row in sd.rows
                     if col_idx - 1 < len(row) and row[col_idx - 1] is not None] + [len(header)],
                    default=10,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        out_path = Path(dest)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return str(out_path)


class GoogleSheetsWriter:
    """Pushes to a Google spreadsheet via gspread. Picks up either a service
    account JSON (production) or OAuth user creds (interactive testing).
    Stub for now — implementation lands when credentials are available."""

    def write(self, sheets: list[SheetData], dest: str) -> str:
        raise NotImplementedError(
            "GoogleSheetsWriter is not yet wired up. Pending service-account credentials. "
            "Use --out-format xlsx for now; the data shape is identical."
        )


def export(out_format: str = "xlsx", out_path: str | None = None,
           spreadsheet_id: str | None = None) -> str:
    """Top-level orchestrator: assemble sheets, pick writer, write.

    For default-folder integration with the briefing pack, use
    `briefing_pack.export()` which calls this with the per-export
    folder's `data.xlsx` path. Standalone CLI usage hits this directly
    via `--export-sheet`."""
    sheets = assemble_sheets()
    if out_format == "xlsx":
        if out_path is None:
            ts = datetime.now().strftime("%Y-%m-%d-%H%M")
            out_path = f"./exports/{ts}/data.xlsx"
        return XlsxWriter().write(sheets, out_path)
    elif out_format == "sheets":
        if not spreadsheet_id:
            raise ValueError("--spreadsheet-id required for --out-format sheets")
        return GoogleSheetsWriter().write(sheets, spreadsheet_id)
    else:
        raise ValueError(f"Unknown out_format: {out_format!r}")
